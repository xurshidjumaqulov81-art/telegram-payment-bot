import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET

import requests
from openpyxl import Workbook, load_workbook
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# =========================
# ENV
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))
CARD_NUMBER = os.getenv("CARD_NUMBER", "")
CARD_OWNER = os.getenv("CARD_OWNER", "")
BASE_DATA_DIR = Path(os.getenv("DATA_DIR", "/app/data"))

# =========================
# PATHS
# =========================
BASE_DATA_DIR.mkdir(parents=True, exist_ok=True)
CHECKS_DIR = BASE_DATA_DIR / "checks"
CHECKS_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = BASE_DATA_DIR / "bot.db"
EXCEL_PATH = BASE_DATA_DIR / "payments.xlsx"

# =========================
# LOGGING
# =========================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# =========================
# STATE NAMES
# =========================
STEP_ID = "id"
STEP_FULL_NAME = "full_name"
STEP_AMOUNT = "amount_usd"
STEP_WAIT_PROOF = "wait_proof"

# =========================
# TEXTS
# =========================
START_TEXT = (
    "👋 Ассалому алайкум!\n\n"
    "🤖 Мен ХЖ тўлов ёрдамчи ботиман.\n\n"
    "Ушбу бот орқали сиз аккаунтингизни тўлдириш учун сўров юборишингиз мумкин.\n\n"
    "📌 Жараённи бошлаш учун қуйидаги тугмани босинг."
)

ASK_ID_TEXT = (
    "🆔 1-қадам: ID рақамингизни киритинг\n\n"
    "Илтимос, аккаунтингизга тегишли ID рақамни киритинг.\n\n"
    "⚠️ ID рақам 7 хонали бўлиши керак.\n\n"
    "Масалан:\n"
    "0012345"
)

ASK_NAME_TEXT = (
    "👤 2-қадам: Исм ва фамилиянгизни киритинг\n\n"
    "Илтимос, аккаунтингиздаги исм ва фамилиянгизни тўлиқ киритинг.\n\n"
    "Масалан:\n"
    "Алиев Алишер"
)

ASK_AMOUNT_TEXT = (
    "💵 3-қадам: Тўлдириш суммасини киритинг\n\n"
    "Илтимос, тўлдирмоқчи бўлган суммани АҚШ долларида киритинг.\n\n"
    "Масалан:\n"
    "10"
)

OFFER_TEXT = (
    "📜 Оммавий оферта шартлари\n\n"
    "Ушбу хизмат орқали амалга оширилган тўловлар фақат аккаунтни тўлдириш мақсадида қабул қилинади.\n\n"
    "⚠️ Фойдаланувчи томонидан киритилган ID рақам ва маълумотлар тўғри бўлиши шарт.\n\n"
    "💳 Тўлов администратор томонидан текширилгандан кейин қабул қилинади.\n\n"
    "⏳ Тўлов тасдиқлангандан кейин маблағ 24 соат ичида аккаунтингизга туширилади."
)

# =========================
# HELPERS
# =========================
NAME_RE = re.compile(r"^[A-Za-zА-Яа-яЁёЎўҚқҒғҲҳʼ'`-]+$")


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_user_id INTEGER NOT NULL,
            telegram_username TEXT,
            account_id TEXT NOT NULL,
            full_name TEXT NOT NULL,
            amount_usd TEXT NOT NULL,
            rate_uzs TEXT NOT NULL,
            amount_uzs INTEGER NOT NULL,
            proof_path TEXT,
            telegram_file_id TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            excel_row INTEGER,
            created_at TEXT NOT NULL,
            approved_at TEXT,
            completed_at TEXT,
            rejected_at TEXT
        )
        """
    )

    conn.commit()
    conn.close()


def init_excel() -> None:
    if EXCEL_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    ws.append(
        [
            "request_id",
            "created_at",
            "telegram_user_id",
            "telegram_username",
            "account_id",
            "full_name",
            "amount_usd",
            "rate_uzs",
            "amount_uzs",
            "proof_path",
            "telegram_file_id",
            "status",
            "approved_at",
            "completed_at",
            "rejected_at",
        ]
    )

    wb.save(EXCEL_PATH)


def append_excel_row(data: dict) -> int:
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.append(
        [
            data["request_id"],
            data["created_at"],
            data["telegram_user_id"],
            data["telegram_username"],
            data["account_id"],
            data["full_name"],
str(data["amount_usd"]),
            data["rate_uzs"],
            data["amount_uzs"],
            data["proof_path"],
            data["telegram_file_id"],
            data["status"],
            data.get("approved_at", ""),
            data.get("completed_at", ""),
            data.get("rejected_at", ""),
        ]
    )

    row_num = ws.max_row
    wb.save(EXCEL_PATH)
    return row_num


def update_excel_status(
    row_num: int,
    *,
    status: str,
    approved_at: str = "",
    completed_at: str = "",
    rejected_at: str = "",
) -> None:
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.cell(row=row_num, column=12).value = status
    if approved_at:
        ws.cell(row=row_num, column=13).value = approved_at
    if completed_at:
        ws.cell(row=row_num, column=14).value = completed_at
    if rejected_at:
        ws.cell(row=row_num, column=15).value = rejected_at

    wb.save(EXCEL_PATH)


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def fmt_uzs(amount: int) -> str:
    return f"{amount:,}".replace(",", " ")


def fmt_decimal_text(value: Decimal) -> str:
    text = f"{value.normalize()}"
    return text.rstrip("0").rstrip(".") if "." in text else text


def parse_amount_to_decimal(text: str) -> Optional[Decimal]:
    cleaned = text.strip().replace(",", ".")
    try:
        value = Decimal(cleaned)
    except Exception:
        return None
    if value <= 0:
        return None
    return value


def validate_full_name(text: str) -> bool:
    parts = text.split()
    if len(parts) < 2:
        return False
    for p in parts:
        if len(p) < 2 or not NAME_RE.match(p):
            return False
    return True


def get_usd_rate_from_cbu() -> Decimal:
    # Official CBU XML endpoint is documented on the webmaster page.
    # Example endpoints are shown there, including currency-specific requests.
    url = "https://cbu.uz/ru/arkhiv-kursov-valyut/xml/USD/"
    response = requests.get(url, timeout=20)
    response.raise_for_status()

    root = ET.fromstring(response.content)
    # Endpoint returns a currency entry; Rate is the exchange rate.
    rate_text = root.findtext(".//Rate")
    if not rate_text:
        raise ValueError("USD kursini olish imkoni bo'lmadi")

    rate = Decimal(rate_text.replace(",", "."))
    return rate


def build_summary_text(account_id: str, full_name: str, amount_usd: Decimal, rate: Decimal, amount_uzs: int) -> str:
    return (
        "📋 Илтимос, маълумотларни текширинг\n\n"
        f"🆔 ID рақам: {account_id}\n"
        f"👤 Исм-фамилия: {full_name}\n"
        f"💵 Сумма: {fmt_decimal_text(amount_usd)} USD\n"
        f"💱 Курс: {rate} сўм\n"
        f"💰 Жами: {fmt_uzs(amount_uzs)} сўм\n\n"
        "Агар маълумотлар тўғри бўлса, тасдиқлаш тугмасини босинг."
    )


def build_admin_text(row: sqlite3.Row) -> str:
    username = f"@{row['telegram_username']}" if row["telegram_username"] else "—"
    return (
        "📥 Янги тўлов сўрови\n\n"
        f"🆔 ID рақам: {row['account_id']}\n"
        f"👤 Исм-фамилия: {row['full_name']}\n"
        f"💵 Сумма: {row['amount_usd']} USD\n"
        f"💱 Курс: {row['rate_uzs']} сўм\n"
        f"💰 Жами: {fmt_uzs(int(row['amount_uzs']))} сўм\n"
        f"👤 Telegram: {username}\n"
        f"🆔 User ID: {row['telegram_user_id']}\n"
        f"🗂 Сўров ID: {row['id']}\n"
        f"🕒 Вақт: {row['created_at']}"
    )


def admin_pending_keyboard(request_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("✅ Тасдиқлаш", callback_data=f"admin_approve:{request_id}"),
                InlineKeyboardButton("❌ Рад этиш", callback_data=f"admin_reject:{request_id}"),
            ]
        ]
    )


def admin_approved_keyboard(request_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("💸 Пул тушди", callback_data=f"admin_complete:{request_id}")]]
    )


# =========================
# DB OPS
# =========================
def create_request_record(
    *,
    telegram_user_id: int,
    telegram_username: str,
    account_id: str,
    full_name: str,
    amount_usd: Decimal,
    rate_uzs: Decimal,
    amount_uzs: int,
    proof_path: str,
    telegram_file_id: str,
) -> int:
    created_at = now_str()

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO requests (
            telegram_user_id,
            telegram_username,
            account_id,
            full_name,
            amount_usd,
            rate_uzs,
            amount_uzs,
            proof_path,
            telegram_file_id,
            status,
            created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        """,
        (
            telegram_user_id,
            telegram_username,
            account_id,
            full_name,
            fmt_decimal_text(amount_usd),
            str(rate_uzs),
            amount_uzs,
            proof_path,
            telegram_file_id,
            created_at,
        ),
    )
    request_id = cur.lastrowid

    excel_row = append_excel_row(
        {
            "request_id": request_id,
            "created_at": created_at,
            "telegram_user_id": telegram_user_id,
            "telegram_username": telegram_username or "",
            "account_id": account_id,
            "full_name": full_name,
            "amount_usd": fmt_decimal_text(amount_usd),
            "rate_uzs": str(rate_uzs),
            "amount_uzs": amount_uzs,
            "proof_path": proof_path,
            "telegram_file_id": telegram_file_id,
            "status": "pending",
        }
    )

    cur.execute("UPDATE requests SET excel_row = ? WHERE id = ?", (excel_row, request_id))
    conn.commit()
    conn.close()

    return request_id


def get_request_by_id(request_id: int) -> Optional[sqlite3.Row]:
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM requests WHERE id = ?", (request_id,))
    row = cur.fetchone()
    conn.close()
    return row


def update_request_status(request_id: int, status: str) -> sqlite3.Row:
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM requests WHERE id = ?", (request_id,))
    row = cur.fetchone()
    if row is None:
        conn.close()
        raise ValueError("Request topilmadi")

    approved_at = row["approved_at"]
    completed_at = row["completed_at"]
    rejected_at = row["rejected_at"]

    current_time = now_str()
    if status == "approved":
        approved_at = current_time
    elif status == "completed":
        completed_at = current_time
    elif status == "rejected":
        rejected_at = current_time

    cur.execute(
        """
        UPDATE requests
        SET status = ?, approved_at = ?, completed_at = ?, rejected_at = ?
        WHERE id = ?
        """,
        (status, approved_at, completed_at, rejected_at, request_id),
    )

    if row["excel_row"]:
        update_excel_status(
            int(row["excel_row"]),
            status=status,
            approved_at=approved_at or "",
            completed_at=completed_at or "",
            rejected_at=rejected_at or "",
        )

    conn.commit()

    cur.execute("SELECT * FROM requests WHERE id = ?", (request_id,))
    updated = cur.fetchone()
    conn.close()
    return updated


# =========================
# BOT HANDLERS
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()

    keyboard = [["🚀 Давом этиш"]]
    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        resize_keyboard=True,
        one_time_keyboard=True,
    )

    if update.message:
        await update.message.reply_text(START_TEXT, reply_markup=reply_markup)


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()
    if update.message:
        await update.message.reply_text(
            "❌ Жараён бекор қилинди.\n\nҚайта бошлаш учун /start ни босинг.",
            reply_markup=ReplyKeyboardRemove(),
        )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.text:
        return

    user_text = update.message.text.strip()
    step = context.user_data.get("step")

    if user_text == "🚀 Давом этиш":
        context.user_data["step"] = STEP_ID
        await update.message.reply_text(ASK_ID_TEXT, reply_markup=ReplyKeyboardRemove())
        return

    if step == STEP_ID:
        if not user_text.isdigit() or len(user_text) != 7:
            await update.message.reply_text(
                "❌ ID рақам нотўғри киритилди.\n\n"
                "Илтимос, 7 хонали ID рақам киритинг.\n\n"
                "Масалан:\n0012345"
            )
            return

        context.user_data["account_id"] = user_text
        context.user_data["step"] = STEP_FULL_NAME
        await update.message.reply_text(ASK_NAME_TEXT)
        return

    if step == STEP_FULL_NAME:
        if not validate_full_name(user_text):
            await update.message.reply_text(
                "❌ Исм ва фамилия нотўғри киритилди.\n\n"
                "Илтимос, тўлиқ ва тўғри киритинг.\n\n"
                "Масалан:\nАлиев Алишер"
            )
            return

        context.user_data["full_name"] = user_text
        context.user_data["step"] = STEP_AMOUNT
        await update.message.reply_text(ASK_AMOUNT_TEXT)
        return

    if step == STEP_AMOUNT:
        amount_usd = parse_amount_to_decimal(user_text)
        if amount_usd is None:
            await update.message.reply_text(
                "❌ Сумма нотўғри киритилди.\n\n"
                "Илтимос, суммани рақам билан киритинг.\n\n"
                "Масалан:\n10"
            )
            return

        try:
            rate = get_usd_rate_from_cbu()
        except Exception as e:
            logger.exception("Kursni olishda xato: %s", e)
            await update.message.reply_text(
                "❌ Ҳозирча курсни олиб бўлмади.\n\n"
                "Бироздан кейин қайта уриниб кўринг."
            )
            return

        amount_uzs = int((amount_usd * rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

        context.user_data["amount_usd"] = str(amount_usd)
        context.user_data["rate_uzs"] = str(rate)
        context.user_data["amount_uzs"] = amount_uzs

        summary_text = build_summary_text(
            account_id=context.user_data["account_id"],
            full_name=context.user_data["full_name"],
            amount_usd=amount_usd,
            rate=rate,
            amount_uzs=amount_uzs,
        )

        keyboard = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ Тасдиқлаш", callback_data="user_confirm"),
                    InlineKeyboardButton("✏️ Ўзгартириш", callback_data="user_restart"),
                ]
            ]
        )

        await update.message.reply_text(summary_text, reply_markup=keyboard)
        return

    if step == STEP_WAIT_PROOF:
        await update.message.reply_text("📸 Илтимос, чек расмини фото кўринишида юборинг.")
        return

    await update.message.reply_text("ℹ️ Жараённи бошлаш учун /start ни босинг.")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.photo:
        return

    step = context.user_data.get("step")
    if step != STEP_WAIT_PROOF:
        await update.message.reply_text("ℹ️ Аввал /start орқали жараённи бошланг.")
        return

    photo = update.message.photo[-1]
    telegram_file_id = photo.file_id

    unique_name = f"{uuid.uuid4().hex}.jpg"
    proof_path = CHECKS_DIR / unique_name

    tg_file = await context.bot.get_file(telegram_file_id)
    await tg_file.download_to_drive(str(proof_path))

    amount_usd = Decimal(context.user_data["amount_usd"])
    rate_uzs = Decimal(context.user_data["rate_uzs"])
    amount_uzs = int(context.user_data["amount_uzs"])

    request_id = create_request_record(
        telegram_user_id=update.effective_user.id,
        telegram_username=update.effective_user.username or "",
        account_id=context.user_data["account_id"],
        full_name=context.user_data["full_name"],
        amount_usd=amount_usd,
        rate_uzs=rate_uzs,
        amount_uzs=amount_uzs,
        proof_path=str(proof_path),
        telegram_file_id=telegram_file_id,
    )

    row = get_request_by_id(request_id)
    assert row is not None

    admin_text = build_admin_text(row)

    await context.bot.send_photo(
        chat_id=ADMIN_ID,
        photo=telegram_file_id,
        caption=admin_text,
        reply_markup=admin_pending_keyboard(request_id),
    )

    await update.message.reply_text(
        "✅ Сўровингиз қабул қилинди\n\n"
        "📤 Маълумотлар администраторга юборилди.\n\n"
        "⏳ Тўлов текширилгандан кейин сизга хабар берилади."
    )

    context.user_data.clear()


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not query:
        return

    await query.answer()
    data = query.data or ""

    # User side
    if data == "user_restart":
        context.user_data.clear()
        context.user_data["step"] = STEP_ID
        await query.message.reply_text(ASK_ID_TEXT)
        await query.edit_message_reply_markup(reply_markup=None)
        return

    if data == "user_confirm":
        keyboard = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ Шартларга розиман", callback_data="offer_accept"),
                    InlineKeyboardButton("❌ Бекор қилиш", callback_data="offer_cancel"),
                ]
            ]
        )
        await query.message.reply_text(OFFER_TEXT, reply_markup=keyboard)
        await query.edit_message_reply_markup(reply_markup=None)
        return

    if data == "offer_cancel":
        context.user_data.clear()
        await query.message.reply_text("❌ Жараён бекор қилинди.\n\nҚайта бошлаш учун /start ни босинг.")
        await query.edit_message_reply_markup(reply_markup=None)
        return

    if data == "offer_accept":
        context.user_data["step"] = STEP_WAIT_PROOF
        amount_uzs = int(context.user_data["amount_uzs"])

        payment_text = (
            "💳 Тўлов учун маълумот\n\n"
            "Қуйидаги карта рақамига тўловни амалга оширинг.\n\n"
            f"💳 Карта:\n{CARD_NUMBER}\n\n"
            f"👤 Қабул қилувчи:\n{CARD_OWNER}\n\n"
            f"💰 Тўлов суммаси:\n{fmt_uzs(amount_uzs)} сўм\n\n"
            "📸 Тўловдан кейин чек расмини юборинг."
        )
        await query.message.reply_text(payment_text)
        await query.edit_message_reply_markup(reply_markup=None)
        return

    # Admin side
    if not data.startswith("admin_"):
        return

    if query.from_user.id != ADMIN_ID:
        await query.answer("Сизда бу амал учун рухсат йўқ.", show_alert=True)
        return

    action, request_id_text = data.split(":")
    request_id = int(request_id_text)

    row = get_request_by_id(request_id)
    if row is None:
        await query.answer("Сўров топилмади.", show_alert=True)
        return

    if action == "admin_approve":
        if row["status"] != "pending":
            await query.answer("Бу сўров аллақачон кўриб чиқилган.", show_alert=True)
            return

        updated = update_request_status(request_id, "approved")

        try:
            await context.bot.send_message(
                chat_id=updated["telegram_user_id"],
                text=(
                    "✅ Тўловингиз тасдиқланди\n\n"
                    "⏳ Маблағ 24 соат ичида аккаунтингизга туширилади."
                ),
            )
        except Exception:
            logger.exception("Userga tasdiq xabarini yuborib bo'lmadi")

        caption = (query.message.caption or "") + "\n\n✅ ТАСДИҚЛАНДИ"
        try:
            await query.edit_message_caption(
                caption=caption,
                reply_markup=admin_approved_keyboard(request_id),
            )
        except Exception:
            await query.message.reply_text(
                "✅ ТАСДИҚЛАНДИ",
                reply_markup=admin_approved_keyboard(request_id),
            )
        return

    if action == "admin_reject":
        if row["status"] != "pending":
            await query.answer("Бу сўровни энди рад этиб бўлмайди.", show_alert=True)
            return

        updated = update_request_status(request_id, "rejected")

        try:
            await context.bot.send_message(
                chat_id=updated["telegram_user_id"],
                text=(
                    "❌ Тўлов сўрови рад этилди.\n\n"
                    "Илтимос, маълумотларни қайта текшириб юборинг ёки администратор билан боғланинг."
                ),
            )
        except Exception:
            logger.exception("Userga rad etish xabarini yuborib bo'lmadi")

        caption = (query.message.caption or "") + "\n\n❌ РАД ЭТИЛДИ"
        try:
            await query.edit_message_caption(caption=caption, reply_markup=None)
        except Exception:
            await query.message.reply_text("❌ РАД ЭТИЛДИ")
        return

    if action == "admin_complete":
        if row["status"] != "approved":
            await query.answer("Аввал тасдиқлаш керак.", show_alert=True)
            return

        updated = update_request_status(request_id, "completed")

        try:
            await context.bot.send_message(
                chat_id=updated["telegram_user_id"],
                text=(
                    "🎉 Муваффақиятли!\n\n"
                    "💰 Маблағ аккаунтингизга туширилди.\n\n"
                    "Хизматдан фойдаланганингиз учун раҳмат."
                ),
            )
        except Exception:
            logger.exception("Userga yakuniy xabarni yuborib bo'lmadi")

        caption = (query.message.caption or "") + "\n\n💸 ПУЛ ТУШДИ"
        try:
            await query.edit_message_caption(caption=caption, reply_markup=None)
        except Exception:
            await query.message.reply_text("💸 ПУЛ ТУШДИ")
        return


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Xatolik yuz berdi", exc_info=context.error)


def validate_env() -> None:
    missing = []
    if not BOT_TOKEN:
        missing.append("BOT_TOKEN")
    if not ADMIN_ID:
        missing.append("ADMIN_ID")
    if not CARD_NUMBER:
        missing.append("CARD_NUMBER")
    if not CARD_OWNER:
        missing.append("CARD_OWNER")

    if missing:
        raise RuntimeError(f"Env topilmadi: {', '.join(missing)}")


def main() -> None:
    validate_env()
    init_db()
    init_excel()

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_error_handler(error_handler)

    logger.info("Bot ishga tushdi")
    app.run_polling()


if __name__ == "__main__":
    main()
